#!/usr/bin/env python3
"""
JM外呼话术质量检查工具
基于人工质检标准，自动化检查话术质量
"""

import openpyxl
import json
import os
import re
from typing import Dict, List, Tuple

# 配置
SCENARIOS_FILE = "/Users/kangrui21/.openclaw/skills/jm-call-script-generator/scenarios_info.json"
OUTPUT_DIR = "/Users/kangrui21/.openclaw/workspace/output_scripts_final"
REPORT_DIR = "/Users/kangrui21/.openclaw/workspace/output_scripts_final/quality_reports"

class QualityChecker:
    """话术质量检查器"""
    
    def __init__(self, scenario_info: Dict):
        self.brand = scenario_info.get("brand", "")
        self.store = scenario_info.get("store", "")
        self.products = scenario_info.get("products", [])
        self.benefits = scenario_info.get("benefits", {})
        self.industry = scenario_info.get("industry", "宠物食品")
        self.pet_type = scenario_info.get("pet_type", "猫")
        self.activity_name = scenario_info.get("activity_name", "")
        self.activity_time = scenario_info.get("activity_time", "")
        
        self.errors = []
        self.warnings = []
        self.checks = []
    
    def check_all(self, wb) -> Tuple[List[str], List[str], List[str]]:
        """执行全部检查"""
        self.errors = []
        self.warnings = []
        self.checks = []
        
        ws_main = wb["主流程"]
        ws_kb = wb["知识库"]
        
        # 1. 品牌名称配置检查
        self._check_brand_config(ws_main, ws_kb)
        
        # 2. 活动时间检查
        self._check_activity_time(ws_main, ws_kb)
        
        # 3. 利益点配置检查
        self._check_benefits_config(ws_main, ws_kb)
        
        # 4. 产品名称检查
        self._check_product_names(ws_main, ws_kb)
        
        # 5. 会员卡级检查
        self._check_member_level(ws_main, ws_kb)
        
        # 6. 利益点叠加检查
        self._check_benefit_overlap(ws_main, ws_kb)
        
        # 7. 客服热线检查
        self._check_customer_service_hotline(ws_main, ws_kb)
        
        # 8. 外呼人设检查
        self._check_caller_identity(ws_main, ws_kb)
        
        # 9. 活动由头检查
        self._check_activity_intro(ws_main, ws_kb)
        
        # 10. 价格规格检查
        self._check_price_spec(ws_main, ws_kb)
        
        # 11. 原模板残留检查
        self._check_template_residual(ws_main, ws_kb)
        
        # 12. 语言通顺性检查
        self._check_language_smoothness(ws_main, ws_kb)
        
        return self.errors, self.warnings, self.checks
    
    def _check_brand_config(self, ws_main, ws_kb):
        """检查品牌名称配置 - 合理标准"""
        check_name = "品牌名称配置"
        
        # 只检查关键节点(开场白、用户忙、重述、邀约)的品牌名称
        key_nodes = ["开场白", "用户忙", "重述3次及以上", "邀约到店-起始节点"]
        
        brand_mentioned = False
        
        # 检查主流程关键节点
        for row in ws_main.iter_rows(min_row=2):
            node_name = str(row[1].value) if row[1].value else ""
            content = str(row[2].value) if row[2].value else ""
            
            if node_name in key_nodes and content and content != "None":
                # 开场白必须包含品牌
                if node_name == "开场白":
                    if self.industry == "日用品":
                        if "京东" not in content:
                            self.errors.append(f"{check_name}: 开场白缺少品牌名称")
                    else:
                        if self.brand and self.brand not in content:
                            self.errors.append(f"{check_name}: 开场白缺少品牌名称 '{self.brand}'")
                    brand_mentioned = True
                
                # 其他关键节点，只需要在至少一个节点提及品牌即可
                if self.brand:
                    if self.brand in content or (self.industry == "日用品" and "京东" in content):
                        brand_mentioned = True
        
        # 如果所有关键节点都没有提及品牌，才报错
        if not brand_mentioned and self.brand:
            self.warnings.append(f"{check_name}: 主流程关键节点未提及品牌名称 '{self.brand}'")
        
        # 检查知识库
        for row in ws_kb.iter_rows(min_row=2):
            question = str(row[0].value) if row[0].value else ""
            answer = str(row[1].value) if row[1].value else ""
            
            if "什么品牌" in question or "什么平台" in question:
                if self.brand and self.brand not in answer:
                    self.errors.append(f"{check_name}: 知识库问答 '{question}' 缺少品牌名称")
            
            if "什么平台" in question and self.store:
                if self.store not in answer:
                    self.errors.append(f"{check_name}: 知识库问答 '{question}' 缺少店铺名称 '{self.store}'")
        
        self.checks.append(f"✓ {check_name}: 已检查品牌 '{self.brand}' 和店铺 '{self.store}' 配置")
    
    def _check_activity_time(self, ws_main, ws_kb):
        """检查活动时间"""
        check_name = "活动时间"
        
        if not self.activity_time:
            self.warnings.append(f"{check_name}: 未配置活动时间")
            return
        
        # 检查主流程是否提及活动时间
        time_mentioned = False
        for row in ws_main.iter_rows(min_row=2):
            content = str(row[2].value) if row[2].value else ""
            if content and self.activity_time in content:
                time_mentioned = True
                break
        
        if not time_mentioned:
            # 如果活动时间较短（如具体日期），应该在关键节点提及
            if re.match(r'\d+\.\d+-\d+\.\d+', self.activity_time):
                self.warnings.append(f"{check_name}: 主流程未提及具体活动时间 '{self.activity_time}'")
        
        self.checks.append(f"✓ {check_name}: 已检查活动时间 '{self.activity_time}'")
    
    def _check_benefits_config(self, ws_main, ws_kb):
        """检查利益点配置"""
        check_name = "利益点配置"
        
        # 提取配置的利益点
        configured_benefits = []
        if "优惠券" in self.benefits:
            configured_benefits.append(self.benefits["优惠券"])
        if "红包" in self.benefits:
            if isinstance(self.benefits["红包"], list):
                configured_benefits.extend(self.benefits["红包"])
        if "大促券" in self.benefits:
            configured_benefits.append(self.benefits["大促券"])
        if "优惠方式" in self.benefits:
            configured_benefits.append(self.benefits["优惠方式"])
        
        # 检查主流程是否提及关键利益点
        main_flow_benefits = []
        for row in ws_main.iter_rows(min_row=2):
            content = str(row[2].value) if row[2].value else ""
            if content and content != "None":
                # 检查是否包含关键利益点关键词
                for benefit in configured_benefits:
                    if benefit and str(benefit) in content:
                        main_flow_benefits.append(benefit)
        
        # 如果配置了利益点但主流程未提及，报错
        for benefit in configured_benefits:
            if benefit and str(benefit) not in main_flow_benefits:
                # 允许部分利益点只在知识库提及
                self.warnings.append(f"{check_name}: 主流程可能未提及利益点 '{benefit}'")
        
        # 检查赠品
        if "赠品" in self.benefits:
            gift_text = self.benefits["赠品"]
            gift_count = self._extract_gift_count(gift_text)
            
            gift_mentioned = False
            for row in ws_main.iter_rows(min_row=2):
                content = str(row[2].value) if row[2].value else ""
                if content and f"{gift_count}件" in content or "好礼" in content:
                    gift_mentioned = True
            
            if not gift_mentioned:
                self.warnings.append(f"{check_name}: 主流程未提及赠品 '{gift_count}件好礼'")
        
        self.checks.append(f"✓ {check_name}: 已检查 {len(configured_benefits)} 个利益点配置")
    
    def _check_product_names(self, ws_main, ws_kb):
        """检查产品名称"""
        check_name = "产品名称"
        
        if not self.products:
            self.warnings.append(f"{check_name}: 未配置产品信息")
            return
        
        # 检查知识库中产品推荐问答
        for row in ws_kb.iter_rows(min_row=2):
            question = str(row[0].value) if row[0].value else ""
            answer = str(row[1].value) if row[1].value else ""
            
            if "推荐什么产品" in question or "卖什么产品" in question:
                # 检查是否包含至少一个产品名称
                product_mentioned = any(product in answer for product in self.products)
                if not product_mentioned:
                    self.errors.append(f"{check_name}: 知识库问答 '{question}' 未提及产品名称")
        
        self.checks.append(f"✓ {check_name}: 已检查 {len(self.products)} 个产品配置")
    
    def _check_member_level(self, ws_main, ws_kb):
        """检查会员卡级/满额加赠"""
        check_name = "会员卡级/满额加赠"
        
        if "加赠" in self.benefits:
            add_gift_text = self.benefits["加赠"]
            
            # 检查主流程是否提及满额加赠
            member_mentioned = False
            for row in ws_main.iter_rows(min_row=2):
                content = str(row[2].value) if row[2].value else ""
                if content and ("满额" in content or "会员" in content or "加赠" in content):
                    member_mentioned = True
            
            if not member_mentioned:
                self.warnings.append(f"{check_name}: 主流程未提及满额加赠条件")
        
        self.checks.append(f"✓ {check_name}: 已检查会员相关权益")
    
    def _check_benefit_overlap(self, ws_main, ws_kb):
        """检查利益点叠加描述"""
        check_name = "利益点叠加描述"
        
        # 统计利益点数量
        benefit_count = 0
        if "优惠券" in self.benefits:
            benefit_count += 1
        if "红包" in self.benefits:
            if isinstance(self.benefits["红包"], list):
                benefit_count += len(self.benefits["红包"])
            else:
                benefit_count += 1
        if "大促券" in self.benefits:
            benefit_count += 1
        if "赠品" in self.benefits:
            benefit_count += 1
        if "加赠" in self.benefits:
            benefit_count += 1
        
        # 如果利益点≥2，检查叠加描述
        if benefit_count >= 2:
            # 检查主流程中是否正确描述叠加关系
            for row in ws_main.iter_rows(min_row=2):
                content = str(row[2].value) if row[2].value else ""
                node_name = str(row[1].value) if row[1].value else ""
                
                if node_name in ["用户忙", "重述3次及以上", "邀约到店-起始节点"]:
                    # 检查叠加关键词
                    if benefit_count >= 2:
                        # 应该有"叠加"、"还有"、"外加"、"满额"等叠加词
                        overlap_keywords = ["叠加", "还有", "外加", "满额", "再送", "加赠"]
                        has_overlap_word = any(kw in content for kw in overlap_keywords)
                        
                        if not has_overlap_word:
                            self.warnings.append(f"{check_name}: 节点 '{node_name}' 有多个利益点但缺少叠加描述词")
        
        self.checks.append(f"✓ {check_name}: 已检查 {benefit_count} 个利益点的叠加描述")
    
    def _check_customer_service_hotline(self, ws_main, ws_kb):
        """检查客服热线"""
        check_name = "客服热线"
        
        # 提取所有客服热线号码
        hotline_pattern = r'(?:客服热线|咨询电话|联系电话)[：:\s]*(\d{3,4}-?\d{7,8})'
        
        hotlines_found = []
        for row in ws_main.iter_rows(min_row=2):
            content = str(row[2].value) if row[2].value else ""
            match = re.search(hotline_pattern, content)
            if match:
                hotlines_found.append(match.group(1))
        
        for row in ws_kb.iter_rows(min_row=2):
            answer = str(row[1].value) if row[1].value else ""
            match = re.search(hotline_pattern, answer)
            if match:
                hotlines_found.append(match.group(1))
        
        if hotlines_found:
            # 如果发现客服热线，需要确认号码正确
            self.warnings.append(f"{check_name}: 发现客服热线号码 {hotlines_found}, 请人工确认号码准确性")
        
        self.checks.append(f"✓ {check_name}: 已检查客服热线（未发现客服热线号码）")
    
    def _check_caller_identity(self, ws_main, ws_kb):
        """检查外呼人设"""
        check_name = "外呼人设"
        
        # 检查开场白中的外呼人设
        for row in ws_main.iter_rows(min_row=2):
            node_name = str(row[1].value) if row[1].value else ""
            content = str(row[2].value) if row[2].value else ""
            
            if node_name == "开场白" and content:
                # 应该包含"这边是XX的"表述
                if "这边是" not in content:
                    self.errors.append(f"{check_name}: 开场白缺少标准外呼人设表述 '这边是XX的'")
                
                # 检查人设名称是否正确
                if self.industry == "日用品":
                    if "京东" not in content:
                        self.errors.append(f"{check_name}: 开场白人设应为 '京东'")
                else:
                    if self.brand and self.brand not in content:
                        self.errors.append(f"{check_name}: 开场白人设应为 '{self.brand}'")
        
        self.checks.append(f"✓ {check_name}: 已检查外呼人设配置")
    
    def _check_activity_intro(self, ws_main, ws_kb):
        """检查活动由头"""
        check_name = "活动由头"
        
        # 检查钩子节点中的活动由头
        hook_row_idx = 7  # 钩子节点通常在第7行
        for row_idx, row in enumerate(ws_main.iter_rows(min_row=2), start=2):
            if row_idx == hook_row_idx:
                content = str(row[2].value) if row[2].value else ""
                if content:
                    # 应该包含"送您XX福利"、"邀请您参加"等由头
                    intro_keywords = ["送您", "邀请您", "给您", "专属福利", "限时"]
                    has_intro = any(kw in content for kw in intro_keywords)
                    
                    if not has_intro:
                        self.warnings.append(f"{check_name}: 钩子节点缺少活动由头表述")
        
        self.checks.append(f"✓ {check_name}: 已检查活动由头表述")
    
    def _check_price_spec(self, ws_main, ws_kb):
        """检查价格规格"""
        check_name = "价格规格"
        
        if "价格" not in self.benefits:
            return
        
        # 检查价格表述是否清晰
        for row in ws_main.iter_rows(min_row=2):
            content = str(row[2].value) if row[2].value else ""
            node_name = str(row[1].value) if row[1].value else ""
            
            if node_name in ["用户忙", "重述3次及以上", "邀约到店-起始节点"]:
                # 检查价格表述是否包含单位
                for product, price_info in self.benefits["价格"].items():
                    if product in content:
                        # 检查是否有价格数字
                        price_pattern = r'\d+(?:\.\d+)?元'
                        if not re.search(price_pattern, content):
                            self.errors.append(f"{check_name}: 节点 '{node_name}' 提及产品 '{product}' 但缺少价格数字")
        
        self.checks.append(f"✓ {check_name}: 已检查价格规格表述")
    
    def _check_template_residual(self, ws_main, ws_kb):
        """检查原模板残留"""
        check_name = "原模板残留"
        
        # 原模板特有内容
        template_residuals = [
            "网易严选紫金烘焙猫粮",
            "紫金烘焙猫粮",
            "83元",
            "9件好礼",
            "猫条",
            "同款试吃",
            "正装宠物香氛",
            "50元专属优惠券",
            "一张50元",
        ]
        
        # 检查主流程
        for row in ws_main.iter_rows(min_row=2):
            content = str(row[2].value) if row[2].value else ""
            node_name = str(row[1].value) if row[1].value else ""
            
            for residual in template_residuals:
                if residual in content:
                    self.errors.append(f"{check_name}: 节点 '{node_name}' 包含原模板残留 '{residual}'")
        
        # 检查知识库
        for row in ws_kb.iter_rows(min_row=2):
            question = str(row[0].value) if row[0].value else ""
            answer = str(row[1].value) if row[1].value else ""
            
            for residual in template_residuals:
                if residual in answer:
                    self.warnings.append(f"{check_name}: 知识库问答 '{question[:20]}' 包含原模板残留 '{residual}'")
        
        self.checks.append(f"✓ {check_name}: 已检查原模板残留")
    
    def _check_language_smoothness(self, ws_main, ws_kb):
        """检查语言通顺性"""
        check_name = "语言通顺性"
        
        # 检查重复表达
        duplicate_patterns = [
            "咱家咱家",
            "券后券后",
            "到手到手",
            "的京东",
            "一张，",
            "低至低至",
        ]
        
        # 检查主流程
        for row in ws_main.iter_rows(min_row=2):
            content = str(row[2].value) if row[2].value else ""
            node_name = str(row[1].value) if row[1].value else ""
            
            for pattern in duplicate_patterns:
                if pattern in content:
                    self.errors.append(f"{check_name}: 节点 '{node_name}' 包含重复表达 '{pattern}'")
        
        # 检查知识库
        for row in ws_kb.iter_rows(min_row=2):
            question = str(row[0].value) if row[0].value else ""
            answer = str(row[1].value) if row[1].value else ""
            
            for pattern in duplicate_patterns:
                if pattern in answer:
                    self.warnings.append(f"{check_name}: 知识库问答 '{question[:20]}' 包含重复表达 '{pattern}'")
        
        self.checks.append(f"✓ {check_name}: 已检查语言通顺性")
    
    def _extract_gift_count(self, gift_text: str) -> str:
        """从赠品文本中提取数量"""
        match = re.search(r'(\d+)件', gift_text)
        if match:
            return match.group(1)
        return gift_text
    
    def generate_report(self, scenario_name: str) -> str:
        """生成质量检查报告"""
        report = f"\n{'='*60}\n"
        report += f"话术质量检查报告 - {scenario_name}\n"
        report += f"{'='*60}\n\n"
        
        # 基本信息
        report += f"【基本信息】\n"
        report += f"  品牌: {self.brand}\n"
        report += f"  店铺: {self.store}\n"
        report += f"  产品: {','.join(self.products) if self.products else '未配置'}\n"
        report += f"  行业: {self.industry}\n"
        report += f"  宠物类型: {self.pet_type if self.industry == '宠物食品' else '无'}\n"
        report += f"  活动名称: {self.activity_name}\n"
        report += f"  活动时间: {self.activity_time if self.activity_time else '未配置'}\n\n"
        
        # 检查结果统计
        report += f"【检查结果统计】\n"
        report += f"  ✅ 检查项: {len(self.checks)} 个\n"
        report += f"  🔴 错误: {len(self.errors)} 个\n"
        report += f"  ⚠️ 警告: {len(self.warnings)} 个\n\n"
        
        # 错误详情
        if self.errors:
            report += f"【错误详情 🔴】\n"
            for i, error in enumerate(self.errors, 1):
                report += f"  {i}. {error}\n"
            report += "\n"
        
        # 警告详情
        if self.warnings:
            report += f"【警告详情 ⚠️】\n"
            for i, warning in enumerate(self.warnings, 1):
                report += f"  {i}. {warning}\n"
            report += "\n"
        
        # 检查项详情
        report += f"【检查项详情 ✅】\n"
        for check in self.checks:
            report += f"  {check}\n"
        report += "\n"
        
        # 质量评分
        total_issues = len(self.errors) + len(self.warnings)
        if total_issues == 0:
            score = "优秀 (A)"
        elif len(self.errors) == 0 and len(self.warnings) <= 2:
            score = "良好 (B)"
        elif len(self.errors) <= 2:
            score = "合格 (C)"
        else:
            score = "需改进 (D)"
        
        report += f"【质量评分】\n"
        report += f"  综合评分: {score}\n"
        report += f"  建议: "
        
        if len(self.errors) > 0:
            report += f"需修复 {len(self.errors)} 个错误后再投入使用\n"
        elif len(self.warnings) > 0:
            report += f"建议优化 {len(self.warnings)} 个警告项以提升话术质量\n"
        else:
            report += f"话术质量优秀，可以直接投入使用\n"
        
        report += f"\n{'='*60}\n"
        
        return report

def check_all_scenarios():
    """检查所有场景的话术质量"""
    # 加载场景信息
    with open(SCENARIOS_FILE, 'r', encoding='utf-8') as f:
        scenarios = json.load(f)
    
    # 创建报告目录
    if not os.path.exists(REPORT_DIR):
        os.makedirs(REPORT_DIR)
    
    print("开始话术质量检查...")
    print(f"检查标准: 基于人工质检要求")
    print(f"检查范围: {len(scenarios)} 个场景\n")
    
    all_reports = []
    
    for scenario_name, scenario_info in scenarios.items():
        print(f"正在检查: {scenario_name}")
        
        # 加载话术文件
        script_file = os.path.join(OUTPUT_DIR, f"{scenario_name}_话术.xlsx")
        if not os.path.exists(script_file):
            print(f"  ⚠️ 文件不存在: {script_file}")
            continue
        
        wb = openpyxl.load_workbook(script_file)
        
        # 创建检查器
        checker = QualityChecker(scenario_info)
        
        # 执行检查
        errors, warnings, checks = checker.check_all(wb)
        
        # 生成报告
        report = checker.generate_report(scenario_name)
        all_reports.append(report)
        
        # 保存报告到文件
        report_file = os.path.join(REPORT_DIR, f"{scenario_name}_质量检查报告.txt")
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report)
        
        print(f"  ✅ 已生成报告: {report_file}")
        print(f"  🔴 错误: {len(errors)} 个")
        print(f"  ⚠️ 警告: {len(warnings)} 个\n")
    
    # 生成汇总报告
    summary_report = generate_summary_report(all_reports)
    summary_file = os.path.join(REPORT_DIR, "话术质量检查汇总报告.txt")
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write(summary_report)
    
    print(f"\n{'='*60}")
    print("质量检查完成!")
    print(f"汇总报告: {summary_file}")
    print(f"报告目录: {REPORT_DIR}")
    print(f"{'='*60}\n")
    
    # 打印汇总报告
    print(summary_report)

def generate_summary_report(all_reports: List[str]) -> str:
    """生成汇总报告"""
    summary = f"\n{'='*80}\n"
    summary += f"话术质量检查汇总报告\n"
    summary += f"{'='*80}\n\n"
    
    summary += f"【检查统计】\n"
    summary += f"  检查场景数: {len(all_reports)} 个\n\n"
    
    # 统计各场景的质量
    summary += f"【各场景质量评估】\n"
    
    total_errors = 0
    total_warnings = 0
    excellent_count = 0
    good_count = 0
    pass_count = 0
    need_improve_count = 0
    
    for report in all_reports:
        # 提取统计信息
        error_count = 0
        warning_count = 0
        score = ""
        
        for line in report.split('\n'):
            if '错误:' in line:
                error_count = int(line.split(':')[1].strip().split()[0])
            elif '警告:' in line:
                warning_count = int(line.split(':')[1].strip().split()[0])
            elif '综合评分:' in line:
                score = line.split(':')[1].strip()
        
        total_errors += error_count
        total_warnings += warning_count
        
        # 分类统计
        if '优秀' in score:
            excellent_count += 1
        elif '良好' in score:
            good_count += 1
        elif '合格' in score:
            pass_count += 1
        else:
            need_improve_count += 1
        
        # 提取场景名称
        scenario_name = ""
        for line in report.split('\n'):
            if '话术质量检查报告 -' in line:
                scenario_name = line.split('-')[1].strip()
                break
        
        summary += f"  {scenario_name}: {score} (错误{error_count}个, 警告{warning_count}个)\n"
    
    summary += f"\n【整体质量分析】\n"
    summary += f"  总错误数: {total_errors} 个\n"
    summary += f"  总警告数: {total_warnings} 个\n"
    summary += f"  优秀(A): {excellent_count} 个场景\n"
    summary += f"  良好(B): {good_count} 个场景\n"
    summary += f"  合格(C): {pass_count} 个场景\n"
    summary += f"  需改进(D): {need_improve_count} 个场景\n\n"
    
    summary += f"【检查标准】\n"
    summary += f"  基于12项人工质检标准:\n"
    summary += f"  1. 品牌名称配置\n"
    summary += f"  2. 活动时间\n"
    summary += f"  3. 利益点配置\n"
    summary += f"  4. 产品名称\n"
    summary += f"  5. 会员卡级/满额加赠\n"
    summary += f"  6. 利益点叠加描述\n"
    summary += f"  7. 客服热线\n"
    summary += f"  8. 外呼人设\n"
    summary += f"  9. 活动由头\n"
    summary += f"  10. 价格规格\n"
    summary += f"  11. 原模板残留\n"
    summary += f"  12. 语言通顺性\n\n"
    
    summary += f"【改进建议】\n"
    if total_errors > 0:
        summary += f"  🔴 需优先修复 {total_errors} 个错误项:\n"
        summary += f"     - 错误项会影响话术的准确性和可用性\n"
        summary += f"     - 建议立即修复后再投入使用\n\n"
    
    if total_warnings > 0:
        summary += f"  ⚠️ 建议优化 {total_warnings} 个警告项:\n"
        summary += f"     - 警告项不影响基本功能，但建议优化以提升话术质量\n"
        summary += f"     - 可以根据实际需求选择性优化\n\n"
    
    if total_errors == 0 and total_warnings == 0:
        summary += f"  ✅ 话术质量优秀:\n"
        summary += f"     - 所有检查项均通过\n"
        summary += f"     - 可以直接投入使用\n\n"
    
    summary += f"{'='*80}\n"
    
    return summary

if __name__ == "__main__":
    check_all_scenarios()