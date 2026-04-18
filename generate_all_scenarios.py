#!/usr/bin/env python3
"""
JM外呼话术生成器 - 多场景批量生成脚本
根据活动信息和模板,为每个场景生成完整话术
"""

import openpyxl
import json
import os
import shutil
from datetime import datetime

# 加载场景信息
SCENARIOS_FILE = "scenarios_info.json"
TEMPLATE_FILE = "/Users/kangrui21/.openclaw/media/outbound/daa9fb1a-4516-4d38-ae25-2bb94c26211a.xlsx"
OUTPUT_DIR = "/Users/kangrui21/.openclaw/workspace/output_scripts"

def load_template():
    """加载原始模板"""
    return openpyxl.load_workbook(TEMPLATE_FILE)

def load_scenarios():
    """加载场景信息"""
    with open(SCENARIOS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def analyze_template_style(ws_main, ws_kb):
    """分析模板语言风格"""
    style = {
        "语气词": [],
        "句式结构": [],
        "称呼": [],
        "表达风格": "亲切口语"
    }
    
    # 从开场白提取语气词
    for row in ws_main.iter_rows(min_row=2, max_row=20):
        content = str(row[2].value) if row[2].value else ""
        if "您好" in content or "喂" in content or "~" in content or "呢" in content:
            if "您好" in content:
                style["语气词"].append("您好")
            if "喂" in content:
                style["语气词"].append("喂")
            if "~" in content:
                style["语气词"].append("~")
    
    # 分析称呼
    if ws_kb:
        for row in ws_kb.iter_rows(min_row=2):
            answer = str(row[1].value) if row[1].value else ""
            if "您" in answer:
                style["称呼"].append("您")
            if "咱家毛孩子" in answer:
                style["称呼"].append("咱家毛孩子")
    
    return style

def generate_sentence(original_sentence, scenario_info, style, node_type):
    """
    根据场景信息生成新句子
    使用完整句子替换策略
    """
    # 场景特定信息
    brand = scenario_info.get("brand", "")
    store = scenario_info.get("store", "")
    products = scenario_info.get("products", [])
    benefits = scenario_info.get("benefits", {})
    industry = scenario_info.get("industry", "宠物食品")
    pet_type = scenario_info.get("pet_type", "猫")
    activity_name = scenario_info.get("activity_name", "限时福利")
    
    # 根据节点类型生成不同内容
    if node_type == "开场白":
        if industry == "日用品":
            return f"喂您好，这边是京东的，您好？（引导用户回应）"
        else:
            return f"喂您好，这边是{brand}的，您好？（引导用户回应）"
    
    elif node_type == "钩子":
        if industry == "日用品":
            return f"就是送您京东专属福利，活动到手价格特别划算，我给您简单说下吧？"
        else:
            return f"就是送您{brand}专属福利，活动到手价格特别划算，我给您简单说下吧？"
    
    elif node_type == "用户忙":
        # 快速说权益
        if industry == "日用品":
            benefit_text = format_benefits_daily(benefits, products)
        elif industry == "宠物食品":
            benefit_text = format_benefits_pet(benefits, products, pet_type)
        else:
            benefit_text = format_benefits_default(benefits, products)
        
        return f"那我快速说下，我们给您京东账户准备了{benefit_text}，待会我把活动链接短信发您，您空了可以点击链接参与选购，感谢您的接听，祝您生活愉快，再见"
    
    elif node_type == "重述3次及以上":
        # 详细权益重述
        if industry == "日用品":
            benefit_text = format_benefits_detail_daily(benefits, products)
        elif industry == "宠物食品":
            benefit_text = format_benefits_detail_pet(benefits, products, pet_type)
        else:
            benefit_text = format_benefits_detail_default(benefits, products)
        
        return f"就是{benefit_text}，待会把活动链接短信发您，您可以点击链接参与选购，感谢您的接听，祝您生活愉快，再见"
    
    elif node_type == "邀约到店":
        # 邀约话术
        if industry == "日用品":
            benefit_text = format_benefits_invite_daily(benefits, products)
        elif industry == "宠物食品":
            benefit_text = format_benefits_invite_pet(benefits, products, pet_type)
        else:
            benefit_text = format_benefits_invite_default(benefits, products)
        
        return f"来电呢是我们给您京东账户准备了{benefit_text}，诚邀您查收下活动短信来享受优惠呀"
    
    elif node_type == "普通节点":
        # 挽回话术
        if industry == "宠物食品":
            if pet_type == "犬":
                return f"啊~那还蛮可惜的，现在活动到手价格特别划算，待会您可以先查收活动短信来店里挑看看嘛~"
            else:
                return f"啊~那还蛮可惜的，现在活动优惠力度很大，待会您可以先查收活动短信给咱家毛孩子挑看看嘛~"
        else:
            return f"啊~那还蛮可惜的，现在活动优惠力度很大，待会您可以先查收活动短信来店里挑看看嘛~"
    
    # 其他节点保持原样或返回空
    return ""

def format_benefits_daily(benefits, products):
    """格式化日用品权益(用户忙节点)"""
    texts = []
    
    # 优惠券
    if "优惠券" in benefits:
        texts.append(f"{benefits['优惠券']}")
    
    # 大促券
    if "大促券" in benefits:
        texts.append(f"还有10周年大促券{benefits['大促券']}")
    
    # 价格
    if "价格" in benefits:
        for product, price in benefits['价格'].items():
            texts.append(f"{product}{price}")
    
    return "、".join(texts)

def format_benefits_pet(benefits, products, pet_type):
    """格式化宠物权益(用户忙节点)"""
    texts = []
    
    # 优惠券
    if "优惠券" in benefits:
        texts.append(f"{benefits['优惠券']}")
    
    # 红包
    if "红包" in benefits:
        red_packets = benefits['红包']
        if isinstance(red_packets, list):
            texts.append(f"还有{','.join(red_packets)}")
    
    # 价格
    if "价格" in benefits:
        for product, price in benefits['价格'].items():
            texts.append(f"{product}券后{price}")
    
    return "、".join(texts)

def format_benefits_default(benefits, products):
    """默认权益格式化"""
    return "专属福利"

def format_benefits_detail_daily(benefits, products):
    """详细权益格式化(重述节点)"""
    texts = []
    
    if "优惠券" in benefits:
        texts.append(f"给您准备了{benefits['优惠券']}")
    
    if "大促券" in benefits:
        texts.append(f"叠加{benefits['大促券']}")
    
    if "价格" in benefits:
        for product, price in benefits['价格'].items():
            texts.append(f"活动到手{product}{price}")
    
    return "、".join(texts)

def format_benefits_detail_pet(benefits, products, pet_type):
    """详细权益格式化(宠物,重述节点)"""
    texts = []
    
    if "优惠券" in benefits:
        texts.append(f"给您准备了{benefits['优惠券']}")
    
    if "红包" in benefits:
        red_packets = benefits['红包']
        if isinstance(red_packets, list):
            texts.append(f"还有{','.join(red_packets)}")
    
    if "价格" in benefits:
        for product, price in benefits['价格'].items():
            texts.append(f"{product}券后{price}")
    
    if "赠品" in benefits:
        texts.append(f"还送{benefits['赠品']}")
    
    return "、".join(texts)

def format_benefits_detail_default(benefits, products):
    return "给您准备了专属福利"

def format_benefits_invite_daily(benefits, products):
    """邀约权益格式化"""
    texts = []
    
    if "优惠券" in benefits:
        texts.append(f"{benefits['优惠券']}")
    
    if "大促券" in benefits:
        texts.append(f"叠加{benefits['大促券']}")
    
    if "价格" in benefits:
        for product, price in benefits['价格'].items():
            texts.append(f"{product}{price}")
    
    return "、".join(texts)

def format_benefits_invite_pet(benefits, products, pet_type):
    """邀约权益格式化(宠物)"""
    texts = []
    
    if "优惠方式" in benefits:
        texts.append(f"{benefits['优惠方式']}")
    elif "优惠券" in benefits:
        texts.append(f"{benefits['优惠券']}")
    
    if "赠品" in benefits:
        texts.append(f"还送{benefits['赠品']}")
    
    return "、".join(texts)

def format_benefits_invite_default(benefits, products):
    return "专属福利"

def update_knowledge_base(ws_kb, scenario_info):
    """更新知识库"""
    brand = scenario_info.get("brand", "")
    store = scenario_info.get("store", "")
    products = scenario_info.get("products", [])
    benefits = scenario_info.get("benefits", {})
    industry = scenario_info.get("industry", "宠物食品")
    pet_type = scenario_info.get("pet_type", "猫")
    activity_name = scenario_info.get("activity_name", "限时福利")
    
    # 需要替换的关键词映射
    replacements = {
        "网易严选紫金烘焙猫粮": brand,
        "网易严选京东自营旗舰店": store,
        "猫粮": products[0] if products else "",
        "83元": str(benefits.get("price", {}).get(products[0], "")),
        "9件好礼": str(benefits.get("赠品", "")),
        "正装宠物香氛": str(benefits.get("加赠", "")),
        "50元专属优惠券": str(benefits.get("优惠券", "")),
        "毛孩子": "咱家狗狗" if pet_type == "犬" else "咱家毛孩子",
        "京东猫粮": f"京东{products[0]}" if industry != "日用品" else "京东",
    }
    
    # 遍历知识库并替换
    for row_idx, row in enumerate(ws_kb.iter_rows(min_row=2), start=2):
        # 问题列(跳过,保持关键词触发)
        # answer列
        if row[1].value:
            answer = str(row[1].value)
            for old_text, new_text in replacements.items():
                if new_text and old_text in answer:
                    # 使用完整句子替换策略
                    answer = answer.replace(old_text, new_text)
            
            # 检查不通顺的地方并修复
            answer = fix_unsmooth_sentences(answer, scenario_info)
            
            ws_kb.cell(row=row_idx, column=2).value = answer
    
    return ws_kb

def fix_unsmooth_sentences(text, scenario_info):
    """修复不通顺的表达"""
    # 移除重复词汇
    if "券后券后" in text:
        text = text.replace("券后券后", "券后")
    
    if "到手到手" in text:
        text = text.replace("到手到手", "到手")
    
    # 修复数量词搭配
    text = text.replace("一张权益", "一份权益")
    text = text.replace("一个活动", "这次活动")
    
    return text

def generate_sms_template(ws_sms, scenario_info):
    """生成短信模板"""
    # 短信模板通常在第二行
    # 这里可以根据场景生成具体短信内容
    # 暂时保持模板结构,内容可以后续手动调整
    return ws_sms

def create_scenario_script(scenario_name, scenario_info, template_wb):
    """为单个场景创建话术文件"""
    # 复制模板
    new_wb = openpyxl.load_workbook(TEMPLATE_FILE)
    
    # 获取各个Sheet
    ws_main = new_wb["主流程"]
    ws_kb = new_wb["知识库"]
    ws_sms = new_wb["短信"]
    
    # 分析模板风格
    style = analyze_template_style(ws_main, ws_kb)
    
    # 更新主流程播报内容
    # 需要识别节点类型并生成对应内容
    for row_idx, row in enumerate(ws_main.iter_rows(min_row=2), start=2):
        node_name = str(row[1].value) if row[1].value else ""
        content = str(row[2].value) if row[2].value else ""
        
        if content and content != "None":
            # 根据节点名称判断类型
            if "开场白" in node_name:
                new_content = generate_sentence(content, scenario_info, style, "开场白")
            elif "钩子" in node_name or row_idx == 7:  # 钩子节点通常在第7行
                new_content = generate_sentence(content, scenario_info, style, "钩子")
            elif "用户忙" in node_name:
                new_content = generate_sentence(content, scenario_info, style, "用户忙")
            elif "重述" in node_name:
                new_content = generate_sentence(content, scenario_info, style, "重述3次及以上")
            elif "邀约" in node_name and "起始" in node_name:
                new_content = generate_sentence(content, scenario_info, style, "邀约到店")
            elif "普通节点" in node_name:
                new_content = generate_sentence(content, scenario_info, style, "普通节点")
            else:
                new_content = content  # 保持原样
            
            # 只更新播报内容列
            if new_content:
                ws_main.cell(row=row_idx, column=3).value = new_content
    
    # 更新知识库
    ws_kb = update_knowledge_base(ws_kb, scenario_info)
    
    # 更新短信模板
    ws_sms = generate_sms_template(ws_sms, scenario_info)
    
    # 保存文件
    output_file = os.path.join(OUTPUT_DIR, f"{scenario_name}_话术.xlsx")
    new_wb.save(output_file)
    
    return output_file

def main():
    """主函数"""
    # 加载场景信息
    scenarios = load_scenarios()
    
    # 加载模板
    template_wb = load_template()
    
    # 创建输出目录
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    print(f"开始生成话术,共{len(scenarios)}个场景...")
    
    # 为每个场景生成话术
    output_files = []
    for scenario_name, scenario_info in scenarios.items():
        print(f"\n正在处理: {scenario_name}")
        output_file = create_scenario_script(scenario_name, scenario_info, template_wb)
        output_files.append(output_file)
        print(f"✓ 已生成: {output_file}")
    
    print(f"\n完成!共生成{len(output_files)}个话术文件")
    print(f"输出目录: {OUTPUT_DIR}")
    
    return output_files

if __name__ == "__main__":
    main()