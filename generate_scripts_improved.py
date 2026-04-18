#!/usr/bin/env python3
"""
JM外呼话术生成器 - 改进版
严格按照SKILL.md的"完整句子替换"和"系统性校验"规范
"""

import openpyxl
import json
import os
import re

# 配置
SCENARIOS_FILE = "scenarios_info.json"
TEMPLATE_FILE = "/Users/kangrui21/.openclaw/media/outbound/daa9fb1a-4516-4d38-ae25-2bb94c26211a.xlsx"
OUTPUT_DIR = "/Users/kangrui21/.openclaw/workspace/output_scripts_improved"

def load_template():
    """加载原始模板"""
    return openpyxl.load_workbook(TEMPLATE_FILE)

def load_scenarios():
    """加载场景信息"""
    with open(SCENARIOS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

class ScriptGenerator:
    """话术生成器"""
    
    def __init__(self, scenario_info):
        self.brand = scenario_info.get("brand", "")
        self.store = scenario_info.get("store", "")
        self.products = scenario_info.get("products", [])
        self.benefits = scenario_info.get("benefits", {})
        self.industry = scenario_info.get("industry", "宠物食品")
        self.pet_type = scenario_info.get("pet_type", "猫")
        self.activity_name = scenario_info.get("activity_name", "限时福利")
        
    def generate开场白(self):
        """生成开场白"""
        if self.industry == "日用品":
            return "喂您好，这边是京东的，您好？（引导用户回应）"
        else:
            return f"喂您好，这边是{self.brand}的，您好？（引导用户回应）"
    
    def generate钩子(self):
        """生成钩子话术"""
        if self.industry == "日用品":
            return "就是送您京东专属福利，活动期间到手价格特别划算，我给您简单说下吧？"
        else:
            return f"就是送您{self.brand}专属福利，活动期间到手价格特别划算，我给您简单说下吧？"
    
    def generate用户忙(self):
        """生成用户忙节点话术(快速说完挂断)"""
        if self.industry == "日用品":
            # 场景一：日用品
            benefits_text = self._format_benefits_daily_quick()
        elif self.industry == "宠物食品":
            # 场景二三四：宠物食品
            benefits_text = self._format_benefits_pet_quick()
        else:
            benefits_text = "专属福利"
        
        return f"那我快速说下，{benefits_text}，待会我把活动链接短信发您，您空了可以点击链接参与选购，感谢您的接听，祝您生活愉快，再见"
    
    def generate重述(self):
        """生成重述节点话术"""
        if self.industry == "日用品":
            benefits_text = self._format_benefits_daily_detail()
        elif self.industry == "宠物食品":
            benefits_text = self._format_benefits_pet_detail()
        else:
            benefits_text = "给您准备了专属福利"
        
        return f"就是{benefits_text}，待会把活动链接短信发您，您可以点击链接参与选购，感谢您的接听，祝您生活愉快，再见"
    
    def generate邀约(self):
        """生成邀约节点话术"""
        if self.industry == "日用品":
            benefits_text = self._format_benefits_daily_invite()
        elif self.industry == "宠物食品":
            benefits_text = self._format_benefits_pet_invite()
        else:
            benefits_text = "专属福利"
        
        return f"来电呢是{benefits_text}，诚邀您查收下活动短信来享受优惠呀"
    
    def generate挽回(self):
        """生成挽回话术"""
        if self.industry == "宠物食品":
            if self.pet_type == "犬":
                return "啊~那还蛮可惜的，现在活动到手价格特别划算，待会您可以先查收活动短信给咱家狗狗挑看看嘛~"
            else:
                return "啊~那还蛮可惜的，现在活动到手价格特别划算，待会您可以先查收活动短信给咱家毛孩子挑看看嘛~"
        else:
            return "啊~那还蛮可惜的，现在活动期间价格特别划算，待会您可以先查收活动短信来店里挑看看嘛~"
    
    def _format_benefits_daily_quick(self):
        """格式化日用品权益(快速版)"""
        texts = []
        
        # 红包
        if "优惠券" in self.benefits:
            texts.append(f"我们给您京东账户准备了{self.benefits['优惠券']}")
        
        # 大促券
        if "大促券" in self.benefits:
            texts.append(f"叠加{self.benefits['大促券']}")
        
        # 价格(更自然的表达)
        if "价格" in self.benefits:
            price_texts = []
            for product, price_info in self.benefits['价格'].items():
                # 解析价格信息: "红包后9.9元(原14.9元)"
                if '红包后' in price_info:
                    price_part = price_info.split('红包后')[1]
                    # 提取价格数字
                    price_num = price_part.split('元')[0]
                    price_texts.append(f"{product}红包后仅{price_num}元")
            if price_texts:
                texts.append(f"活动到手{','.join(price_texts)}")
        
        return "、".join(texts)
    
    def _format_benefits_daily_detail(self):
        """格式化日用品权益(详细版)"""
        texts = []
        
        if "优惠券" in self.benefits:
            texts.append(f"给您准备了{self.benefits['优惠券']}")
        
        if "大促券" in self.benefits:
            texts.append(f"叠加{self.benefits['大促券']}")
        
        if "价格" in self.benefits:
            price_texts = []
            for product, price_info in self.benefits['价格'].items():
                if '红包后' in price_info:
                    price_part = price_info.split('红包后')[1]
                    price_num = price_part.split('元')[0]
                    price_texts.append(f"{product}红包后仅{price_num}元")
            if price_texts:
                texts.append(f"活动到手{','.join(price_texts)}")
        
        return "、".join(texts)
    
    def _format_benefits_daily_invite(self):
        """格式化日用品权益(邀约版)"""
        texts = []
        
        if "优惠券" in self.benefits:
            texts.append(f"我们给您京东账户准备了{self.benefits['优惠券']}")
        
        if "大促券" in self.benefits:
            texts.append(f"叠加{self.benefits['大促券']}")
        
        if "价格" in self.benefits:
            price_texts = []
            for product, price_info in self.benefits['价格'].items():
                if '红包后' in price_info:
                    price_part = price_info.split('红包后')[1]
                    price_num = price_part.split('元')[0]
                    price_texts.append(f"{product}红包后仅{price_num}元")
            if price_texts:
                texts.append(f"活动到手{','.join(price_texts)}")
        
        return "、".join(texts)
    
    def _format_benefits_pet_quick(self):
        """格式化宠物权益(快速版)"""
        texts = []
        
        # 场景二：有8折券+红包
        if "优惠券" in self.benefits:
            texts.append(f"我们给您准备了{self.benefits['优惠券']}")
        
        if "红包" in self.benefits:
            red_packets = self.benefits['红包']
            if isinstance(red_packets, list):
                texts.append(f"还有{','.join(red_packets)}")
        
        # 场景三四：叠券到手价
        if "优惠方式" in self.benefits:
            # 简化表达: "买4袋叠券后到手476元" -> "买4袋叠券后476元"
            benefit_text = self.benefits['优惠方式']
            if '到手' in benefit_text:
                benefit_text = benefit_text.replace('到手', '')
            texts.append(benefit_text)
        
        # 价格信息(场景二专用)
        if "价格" in self.benefits:
            for product, price in self.benefits['价格'].items():
                # 简化价格表达: "8折券后271.2元(台前价339,1.8kg*4袋)" -> "271.2元"
                if '券后' in price:
                    # 提取券后价格
                    price_parts = price.split('券后')
                    if len(price_parts) > 1:
                        price_num = price_parts[1].split('元')[0].replace('台前价', '').replace('8折', '')
                        texts.append(f"{product}券后{price_num}元")
        
        # 赠品
        if "赠品" in self.benefits:
            gift_count = self._extract_gift_count(self.benefits['赠品'])
            texts.append(f"还送{gift_count}件好礼")
        
        # 加赠
        if "加赠" in self.benefits:
            # 简化表达
            add_gift_text = self.benefits['加赠']
            if '宠物单笔实付满239元加赠' in add_gift_text:
                add_gift_text = add_gift_text.replace('宠物单笔实付满239元加赠', '满239元加赠')
            texts.append(add_gift_text)
        
        return "、".join(texts)
    
    def _format_benefits_pet_detail(self):
        """格式化宠物权益(详细版)"""
        # 与快速版类似,但表达更详细
        return self._format_benefits_pet_quick()
    
    def _format_benefits_pet_invite(self):
        """格式化宠物权益(邀约版)"""
        texts = []
        
        # 优惠方式
        if "优惠方式" in self.benefits:
            texts.append(f"给您准备了{self.benefits['优惠方式']}")
        elif "优惠券" in self.benefits:
            texts.append(f"给您准备了{self.benefits['优惠券']}")
        
        # 赠品
        if "赠品" in self.benefits:
            gift_count = self._extract_gift_count(self.benefits['赠品'])
            texts.append(f"还送{gift_count}件好礼")
        
        # 加赠
        if "加赠" in self.benefits:
            texts.append(f"满额再加赠{self.benefits['加赠']}")
        
        return "、".join(texts)
    
    def _extract_gift_count(self, gift_text):
        """从赠品文本中提取数量"""
        # 提取数字
        match = re.search(r'(\d+)件', gift_text)
        if match:
            return match.group(1)
        return gift_text
    
    def update_knowledge_base(self, ws_kb):
        """更新知识库 - 完整句子替换"""
        # 需要完全替换的内容
        full_replacements = {
            "语音助手": self._generate_kb_voice_assistant(),
            "推荐什么产品": self._generate_kb_product_recommend(),
            "有什么活动/福利": self._generate_kb_benefits(),
            "咨询价格": self._generate_kb_price(),
            "奖品/礼品/礼包/礼盒是什么": self._generate_kb_gift(),
            "来电意图": self._generate_kb_intent(),
            "什么品牌的来电": self._generate_kb_brand(),
            "什么平台": self._generate_kb_platform(),
            "是线上还是线下活动": self._generate_kb_channel(),
            "你是机器人吗": self._generate_kb_robot(),
            "为什么是外地号码": self._generate_kb_number(),
            "活动不给力": self._generate_kb_activity_strength(),
            "还有其他活动/福利吗": self._generate_kb_other_benefits(),
        }
        
        # 遍历知识库并替换
        for row_idx, row in enumerate(ws_kb.iter_rows(min_row=2), start=2):
            question = str(row[0].value) if row[0].value else ""
            
            if question in full_replacements:
                # 完整替换回答
                ws_kb.cell(row=row_idx, column=2).value = full_replacements[question]
            else:
                # 其他问题,检查是否有原模板残留并清理
                answer = str(row[1].value) if row[1].value else ""
                answer = self._clean_template_residual(answer)
                ws_kb.cell(row=row_idx, column=2).value = answer
        
        return ws_kb
    
    def _generate_kb_voice_assistant(self):
        """生成语音助手问答"""
        if self.industry == "日用品":
            benefits = self._format_benefits_daily_quick()
            return f"您好，我是京东的，来电呢是送您{benefits}，待会给您发个短信链接，趁活动优惠快来看看吧。感谢接听，祝您生活愉快，再见。"
        else:
            benefits = self._format_benefits_pet_quick()
            pet_call = "咱家狗狗" if self.pet_type == "犬" else "咱家毛孩子"
            return f"您好，我是{self.brand}的，来电呢是送您{benefits}，待会给您发个短信链接，趁活动优惠快给{pet_call}囤点好吃的吧。感谢接听，祝您生活愉快，再见。"
    
    def _generate_kb_product_recommend(self):
        """生成产品推荐问答"""
        if self.products:
            product_names = ",".join(self.products[:2])  #最多列出2个产品
            if self.industry == "宠物食品":
                pet_call = "咱家狗狗" if self.pet_type == "犬" else "咱家毛孩子"
                return f"这次给您推荐的是{product_names}，具体的可以根据{pet_call}的体重、需求，找店铺客服做推荐哦"
            else:
                return f"这次给您推荐的是{product_names}，具体的可以根据您的需求，找店铺客服做推荐哦"
        return "这次给您推荐的是我们店铺的热销产品，具体建议您进入店铺了解哦"
    
    def _generate_kb_benefits(self):
        """生成活动福利问答"""
        if self.industry == "日用品":
            benefits = self._format_benefits_daily_detail()
        else:
            benefits = self._format_benefits_pet_detail()
        
        return f"我们{benefits}，待会给您发个短信链接，您点击链接就可以参与活动了哦"
    
    def _generate_kb_price(self):
        """生成价格咨询问答"""
        if "价格" in self.benefits:
            price_texts = []
            for product, price in self.benefits['价格'].items():
                if self.industry == "日用品":
                    # 清理价格表达
                    price_clean = price.split('红包后')[1].split('(')[0] if '红包后' in price else price
                    price_texts.append(f"{product}红包后{price_clean}")
                else:
                    price_texts.append(f"{product}活动价{price}")
            return f"像刚刚给您推荐的产品，活动到手价是{','.join(price_texts)}，具体价格看您最后选购的商品了哈，您可以在活动页面下单时咨询客服了解哦"
        return "活动期间多款产品都有超值价格，具体建议您进店咨询客服了解哦"
    
    def _generate_kb_gift(self):
        """生成赠品问答"""
        if "赠品" in self.benefits:
            gift_count = self._extract_gift_count(self.benefits['赠品'])
            gift_detail = self.benefits['赠品']
            
            if "加赠" in self.benefits:
                return f"下单就送{gift_count}件好礼，有{gift_detail}，满额还额外{self.benefits['加赠']},那更多礼赠详情建议您加购看一下商品详情页面或者咨询客服啊"
            else:
                return f"下单就送{gift_count}件好礼，有{gift_detail},那更多礼赠详情建议您加购看一下商品详情页面或者咨询客服啊"
        return "活动期间下单有超值赠品，具体建议您进店咨询客服了解哦"
    
    def _generate_kb_intent(self):
        """生成来电意图问答"""
        if self.industry == "日用品":
            benefits = self._format_benefits_daily_quick()
            return f"打电话是{benefits}，待会给您发个短信链接，您点击链接就可以参与活动了哦"
        else:
            benefits = self._format_benefits_pet_quick()
            pet_call = "咱家狗狗" if self.pet_type == "犬" else "咱家毛孩子"
            return f"打电话是{benefits}，待会给您发个短信链接，趁活动优惠快给{pet_call}囤点好吃的吧"
    
    def _generate_kb_brand(self):
        """生成品牌问答"""
        return f"我这边是{self.brand}的，这次来电是邀请您参加我们{self.store}限时优惠活动的哈~"
    
    def _generate_kb_platform(self):
        """生成平台问答"""
        return f"邀请您参加的是【京东平台】【{self.store}】的专享活动哈"
    
    def _generate_kb_channel(self):
        """生成渠道问答"""
        return f"本次邀请您参加的是【京东APP】【{self.store}】的活动哦，稍后您查收短信看一看嘛"
    
    def _generate_kb_robot(self):
        """生成机器人问答"""
        return f"哈哈，还是被您发现啦，我是{self.brand}的智能语音小助手，来电主要是给您送{self.store}限时活动福利的呢"
    
    def _generate_kb_number(self):
        """生成外地号码问答"""
        return f"因为使用的是网络拨号，号码是运营商随机显示的，请您放心，我是{self.brand}官方来电~来电也是给您分享{self.store}限时活动福利的"
    
    def _generate_kb_activity_strength(self):
        """生成活动力度问答"""
        return f"我们这次是{self.brand}限时专享活动，给到您的也是很优惠的价格了，建议您可以到我们{self.store}对比看下哦"
    
    def _generate_kb_other_benefits(self):
        """生成其他福利问答"""
        benefits_text = ""
        if "加赠" in self.benefits:
            benefits_text = f"，满额还{self.benefits['加赠']}"
        
        return f"除了优惠价格和礼赠{benefits_text}，那更多活动，您可以到{self.store}联系店铺客服详细了解一下呢~"
    
    def _clean_template_residual(self, text):
        """清理原模板残留"""
        # 删除原模板特有内容
        residuals_to_remove = [
            "网易严选紫金烘焙猫粮",
            "83元",
            "9件好礼",
            "猫条",
            "同款试吃",
            "正装宠物香氛",
            "50元专属优惠券",
            "紫金烘焙猫粮",
            "猫条等9件好礼",
            "同款试吃装",
        ]
        
        for residual in residuals_to_remove:
            if residual in text:
                text = text.replace(residual, "")
        
        # 修复双重称呼
        text = text.replace("咱家咱家", "咱家")
        text = text.replace("券后券后", "券后")
        text = text.replace("到手到手", "到手")
        
        # 修复不通顺的表达
        text = text.replace("一张，", "一张")
        text = text.replace("低至83元每包", "")
        text = text.replace("低至83元每袋", "")
        
        return text

def create_improved_script(scenario_name, scenario_info, template_wb):
    """创建改进版话术文件"""
    # 复制模板
    new_wb = openpyxl.load_workbook(TEMPLATE_FILE)
    
    ws_main = new_wb["主流程"]
    ws_kb = new_wb["知识库"]
    ws_sms = new_wb["短信"]
    
    # 创建生成器
    generator = ScriptGenerator(scenario_info)
    
    # 更新主流程
    for row_idx, row in enumerate(ws_main.iter_rows(min_row=2), start=2):
        node_name = str(row[1].value) if row[1].value else ""
        
        # 根据节点名称生成内容
        if node_name == "开场白":
            ws_main.cell(row=row_idx, column=3).value = generator.generate开场白()
        elif row_idx == 7:  # 钩子节点
            ws_main.cell(row=row_idx, column=3).value = generator.generate钩子()
        elif node_name == "用户忙":
            ws_main.cell(row=row_idx, column=3).value = generator.generate用户忙()
        elif "重述" in node_name:
            ws_main.cell(row=row_idx, column=3).value = generator.generate重述()
        elif node_name == "邀约到店-起始节点":
            ws_main.cell(row=row_idx, column=3).value = generator.generate邀约()
        elif node_name == "普通节点":
            ws_main.cell(row=row_idx, column=3).value = generator.generate挽回()
    
    # 更新知识库
    # 更新知识库时也需要处理特殊情况
        # 检查是否有空的产品列表
        if not self.products or len(self.products) == 0:
            ws_kb.cell(row=row_idx, column=2).value = "这次给您推荐的是我们店铺的热销产品，具体建议您进入店铺了解哦"
        else:
            ws_kb = generator.update_knowledge_base(ws_kb)
    
    # 保存文件
    output_file = os.path.join(OUTPUT_DIR, f"{scenario_name}_话术.xlsx")
    new_wb.save(output_file)
    
    return output_file

def main():
    """主函数"""
    scenarios = load_scenarios()
    
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    print(f"开始生成改进版话术,共{len(scenarios)}个场景...")
    
    output_files = []
    for scenario_name, scenario_info in scenarios.items():
        print(f"\n正在处理: {scenario_name}")
        template_wb = load_template()
        output_file = create_improved_script(scenario_name, scenario_info, template_wb)
        output_files.append(output_file)
        print(f"✓ 已生成: {output_file}")
    
    print(f"\n完成!共生成{len(output_files)}个话术文件")
    print(f"输出目录: {OUTPUT_DIR}")

if __name__ == "__main__":
    main()