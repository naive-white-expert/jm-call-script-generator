#!/usr/bin/env python3
"""
JM外呼话术生成器 - 最终优化版
基于质量检查结果,修复所有错误项
"""

import openpyxl
import json
import os
import re

# 配置
SCENARIOS_FILE = "/Users/kangrui21/.openclaw/skills/jm-call-script-generator/scenarios_info.json"
TEMPLATE_FILE = "/Users/kangrui21/.openclaw/media/outbound/daa9fb1a-4516-4d38-ae25-2bb94c26211a.xlsx"
OUTPUT_DIR = "/Users/kangrui21/.openclaw/workspace/output_scripts_final"

class FinalScriptGenerator:
    """最终版话术生成器 - 修复所有质量问题"""
    
    def __init__(self, scenario_info):
        self.brand = scenario_info.get("brand", "")
        self.store = scenario_info.get("store", "")
        self.products = scenario_info.get("products", [])
        self.benefits = scenario_info.get("benefits", {})
        self.industry = scenario_info.get("industry", "宠物食品")
        self.pet_type = scenario_info.get("pet_type", "猫")
        self.activity_name = scenario_info.get("activity_name", "限时福利")
        self.activity_time = scenario_info.get("activity_time", "")
    
    def generate开场白(self):
        """开场白 - 必须包含品牌名称"""
        if self.industry == "日用品":
            return "喂您好，这边是京东的，您好？（引导用户回应）"
        else:
            return f"喂您好，这边是{self.brand}的，您好？（引导用户回应）"
    
    def generate钩子(self):
        """钩子 - 提及活动时间(如果有具体时间)"""
        time_text = ""
        if self.activity_time and re.match(r'\d+\.\d+-\d+\.\d+', self.activity_time):
            time_text = f"，活动时间{self.activity_time}"
        
        if self.industry == "日用品":
            return f"就是送您京东{self.activity_name}专属福利{time_text}，活动期间到手价格特别划算，我给您简单说下吧？"
        else:
            return f"就是送您{self.brand}{self.activity_name}专属福利{time_text}，活动期间到手价格特别划算，我给您简单说下吧？"
    
    def generate用户忙(self):
        """用户忙节点 - 快速说完挂断"""
        benefits_text = self._format_benefits_complete()
        return f"那我快速说下，{benefits_text}，待会我把活动链接短信发您，您空了可以点击链接参与选购，感谢您的接听，祝您生活愉快，再见"
    
    def generate重述(self):
        """重述节点 - 详细权益说明"""
        benefits_text = self._format_benefits_complete()
        return f"就是{benefits_text}，待会把活动链接短信发您，您可以点击链接参与选购，感谢您的接听，祝您生活愉快，再见"
    
    def generate邀约(self):
        """邀约节点 - 邀约到店"""
        benefits_text = self._format_benefits_invite()
        return f"来电呢是{benefits_text}，诚邀您查收下活动短信来享受优惠呀"
    
    def generate挽回(self):
        """挽回节点 - 挽回话术"""
        if self.industry == "宠物食品":
            pet_call = "咱家狗狗" if self.pet_type == "犬" else "咱家毛孩子"
            return f"啊~那还蛮可惜的，现在活动到手价格特别划算，待会您可以先查收活动短信给{pet_call}挑看看嘛~"
        else:
            return "啊~那还蛮可惜的，现在活动期间价格特别划算，待会您可以先查收活动短信来店里挑看看嘛~"
    
    def _format_benefits_complete(self):
        """格式化完整权益信息(用户忙+重述节点)"""
        texts = []
        
        # 根据行业类型格式化
        if self.industry == "日用品":
            # 场景一：日用品
            if "优惠券" in self.benefits:
                texts.append(f"我们给您京东账户准备了{self.benefits['优惠券']}")
            
            if "大促券" in self.benefits:
                texts.append(f"叠加{self.benefits['大促券']}")
            
            if "价格" in self.benefits:
                price_parts = []
                for product, price_info in self.benefits['价格'].items():
                    if '红包后' in price_info:
                        price_num = price_info.split('红包后')[1].split('元')[0]
                        price_parts.append(f"{product}红包后仅{price_num}元")
                if price_parts:
                    texts.append(f"活动到手{','.join(price_parts)}")
        
        elif self.industry == "宠物食品":
            # 场景二三四：宠物食品
            # 优惠券/优惠方式
            if "优惠方式" in self.benefits:
                benefit_text = self.benefits['优惠方式'].replace('到手', '')
                texts.append(f"我们给您准备了{benefit_text}")
            elif "优惠券" in self.benefits:
                texts.append(f"我们给您准备了{self.benefits['优惠券']}")
            
            # 红包
            if "红包" in self.benefits:
                red_packets = self.benefits['红包']
                if isinstance(red_packets, list):
                    texts.append(f"还有{','.join(red_packets)}")
            
            # 价格(场景二专用)
            if "价格" in self.benefits:
                for product, price in self.benefits['价格'].items():
                    if '券后' in price:
                        price_parts = price.split('券后')
                        if len(price_parts) > 1:
                            price_num = price_parts[1].split('元')[0].replace('台前价', '').replace('8折', '')
                            texts.append(f"{product}券后{price_num}元")
            
            # 赠品
            if "赠品" in self.benefits:
                gift_count = self._extract_gift_count(self.benefits['赠品'])
                texts.append(f"还送{gift_count}件好礼")
            
            # 加赠
            if "加赠" in self.benefits:
                add_gift_text = self.benefits['加赠']
                if '宠物单笔实付满239元加赠' in add_gift_text:
                    add_gift_text = add_gift_text.replace('宠物单笔实付满239元加赠', '满239元还')
                texts.append(add_gift_text)
        
        return "、".join(texts)
    
    def _format_benefits_invite(self):
        """格式化邀约权益(邀约节点)"""
        texts = []
        
        if self.industry == "日用品":
            if "优惠券" in self.benefits:
                texts.append(f"给您京东账户准备了{self.benefits['优惠券']}")
            
            if "大促券" in self.benefits:
                texts.append(f"叠加{self.benefits['大促券']}")
            
            if "价格" in self.benefits:
                price_parts = []
                for product, price_info in self.benefits['价格'].items():
                    if '红包后' in price_info:
                        price_num = price_info.split('红包后')[1].split('元')[0]
                        price_parts.append(f"{product}红包后仅{price_num}元")
                if price_parts:
                    texts.append(f"活动到手{','.join(price_parts)}")
        
        else:
            # 宠物食品
            if "优惠方式" in self.benefits:
                texts.append(f"给您准备了{self.benefits['优惠方式'].replace('到手', '')}")
            elif "优惠券" in self.benefits:
                texts.append(f"给您准备了{self.benefits['优惠券']}")
            
            if "赠品" in self.benefits:
                gift_count = self._extract_gift_count(self.benefits['赠品'])
                texts.append(f"还送{gift_count}件好礼")
            
            if "加赠" in self.benefits:
                add_gift_text = self.benefits['加赠']
                if '宠物单笔实付满239元加赠' in add_gift_text:
                    add_gift_text = add_gift_text.replace('宠物单笔实付满239元加赠', '满239元还')
                texts.append(add_gift_text)
        
        return "、".join(texts)
    
    def update_knowledge_base(self, ws_kb):
        """更新知识库 - 完全修复所有问题"""
        # 完整替换的知识库问答
        full_replacements = {
            "语音助手": self._generate_kb_voice_assistant(),
            "推荐什么产品": self._generate_kb_product_recommend(),
            "有什么活动/福利": self._generate_kb_benefits(),
            "咨询价格": self._generate_kb_price(),
            "奖品/礼品/礼包/礼盒是什么": self._generate_kb_gift(),
            "来电意图": self._generate_kb_intent(),
            "卖什么产品": self._generate_kb_sell_product(),  # 新增：修复产品名称错误
            "什么品牌的来电": self._generate_kb_brand(),
            "什么平台": self._generate_kb_platform(),
            "是线上还是线下活动": self._generate_kb_channel(),
            "你是机器人吗": self._generate_kb_robot(),
            "为什么是外地号码": self._generate_kb_number(),
            "活动不给力": self._generate_kb_activity_strength(),
            "还有其他活动/福利吗": self._generate_kb_other_benefits(),
            "怎么满减": self._generate_kb_discount_rule(),  # 新增：修复原模板残留
            "号码哪里来的": self._generate_kb_number_source(),  # 新增：修复重复表达
        }
        
        # 遍历知识库并替换
        for row_idx, row in enumerate(ws_kb.iter_rows(min_row=2), start=2):
            question = str(row[0].value) if row[0].value else ""
            
            if question in full_replacements:
                ws_kb.cell(row=row_idx, column=2).value = full_replacements[question]
            else:
                # 其他问答，清理残留
                answer = str(row[1].value) if row[1].value else ""
                answer = self._clean_all_residuals(answer)
                ws_kb.cell(row=row_idx, column=2).value = answer
        
        return ws_kb
    
    def _generate_kb_sell_product(self):
        """生成'卖什么产品'问答 - 修复产品名称错误"""
        if self.products:
            product_names = ",".join(self.products[:2])
            return f"这次给您推荐的是{product_names}的福利，具体建议您进入店铺了解，或者在活动页面咨询客服哦"
        return "这次给您推荐的是我们店铺的热销产品，具体建议您进入店铺了解哦"
    
    def _generate_kb_discount_rule(self):
        """生成'怎么满减'问答 - 修复原模板残留"""
        if "优惠券" in self.benefits:
            return f"送您的是{self.benefits['优惠券']}，那具体满减的话您可以进店咨询客服了解哦"
        elif "优惠方式" in self.benefits:
            return f"{self.benefits['优惠方式']}，那具体规则的话您可以进店咨询客服了解哦"
        return "具体的优惠规则您可以进店咨询客服了解哦"
    
    def _generate_kb_number_source(self):
        """生成'号码哪里来的'问答 - 修复重复表达"""
        return "因为您是我们尊贵的京东用户嘛，所以有您的联系方式，不过您放心信息都是加密的"
    
    def _generate_kb_voice_assistant(self):
        """语音助手问答"""
        if self.industry == "日用品":
            benefits = self._format_benefits_complete()
            return f"您好，我是京东的，来电呢是{benefits}，待会给您发个短信链接，趁活动优惠快来看看吧。感谢接听，祝您生活愉快，再见。"
        else:
            benefits = self._format_benefits_complete()
            pet_call = "咱家狗狗" if self.pet_type == "犬" else "咱家毛孩子"
            return f"您好，我是{self.brand}的，来电呢是{benefits}，待会给您发个短信链接，趁活动优惠快给{pet_call}囤点好吃的吧。感谢接听，祝您生活愉快，再见。"
    
    def _generate_kb_product_recommend(self):
        """产品推荐问答"""
        if self.products:
            product_names = ",".join(self.products[:2])
            if self.industry == "宠物食品":
                pet_call = "咱家狗狗" if self.pet_type == "犬" else "咱家毛孩子"
                return f"这次给您推荐的是{product_names}，具体的可以根据{pet_call}的体重、需求，找店铺客服做推荐哦"
            else:
                return f"这次给您推荐的是{product_names}，具体的可以根据您的需求，找店铺客服做推荐哦"
        return "这次给您推荐的是我们店铺的热销产品，具体建议您进入店铺了解哦"
    
    def _generate_kb_benefits(self):
        """活动福利问答"""
        benefits = self._format_benefits_complete()
        return f"我们{benefits}，待会给您发个短信链接，您点击链接就可以参与活动了哦"
    
    def _generate_kb_price(self):
        """价格咨询问答"""
        if "价格" in self.benefits:
            price_texts = []
            for product, price in self.benefits['价格'].items():
                if self.industry == "日用品":
                    if '红包后' in price:
                        price_num = price.split('红包后')[1].split('元')[0]
                        price_texts.append(f"{product}红包后{price_num}元")
                else:
                    if '券后' in price:
                        price_num = price.split('券后')[1].split('元')[0].replace('台前价', '').replace('8折', '')
                        price_texts.append(f"{product}券后{price_num}元")
            if price_texts:
                return f"像刚刚给您推荐的产品，活动到手价是{','.join(price_texts)}，具体价格看您最后选购的商品了哈，您可以在活动页面下单时咨询客服了解哦"
        return "活动期间多款产品都有超值价格，具体建议您进店咨询客服了解哦"
    
    def _generate_kb_gift(self):
        """赠品问答"""
        if "赠品" in self.benefits:
            gift_count = self._extract_gift_count(self.benefits['赠品'])
            gift_detail = self.benefits['赠品']
            
            if "加赠" in self.benefits:
                add_gift_text = self.benefits['加赠']
                if '宠物单笔实付满239元加赠' in add_gift_text:
                    add_gift_text = add_gift_text.replace('宠物单笔实付满239元加赠', '满239元还')
                return f"下单就送{gift_count}件好礼，有{gift_detail}，{add_gift_text}，那更多礼赠详情建议您加购看一下商品详情页面或者咨询客服啊"
            else:
                return f"下单就送{gift_count}件好礼，有{gift_detail}，那更多礼赠详情建议您加购看一下商品详情页面或者咨询客服啊"
        return "活动期间下单有超值赠品，具体建议您进店咨询客服了解哦"
    
    def _generate_kb_intent(self):
        """来电意图问答"""
        benefits = self._format_benefits_invite()
        if self.industry == "宠物食品":
            pet_call = "咱家狗狗" if self.pet_type == "犬" else "咱家毛孩子"
            return f"打电话是{benefits}，待会给您发个短信链接，趁活动优惠快给{pet_call}囤点好吃的吧"
        else:
            return f"打电话是{benefits}，待会给您发个短信链接，您点击链接就可以参与活动了哦"
    
    def _generate_kb_brand(self):
        """品牌问答"""
        return f"我这边是{self.brand}的，这次来电是邀请您参加我们{self.store}限时优惠活动的哈~"
    
    def _generate_kb_platform(self):
        """平台问答"""
        return f"邀请您参加的是【京东平台】【{self.store}】的专享活动哈"
    
    def _generate_kb_channel(self):
        """渠道问答"""
        return f"本次邀请您参加的是【京东APP】【{self.store}】的活动哦，稍后您查收短信看一看嘛"
    
    def _generate_kb_robot(self):
        """机器人问答"""
        return f"哈哈，还是被您发现啦，我是{self.brand}的智能语音小助手，来电主要是给您送{self.store}限时活动福利的呢"
    
    def _generate_kb_number(self):
        """外地号码问答"""
        return f"因为使用的是网络拨号，号码是运营商随机显示的，请您放心，我是{self.brand}官方来电~来电也是给您分享{self.store}限时活动福利的"
    
    def _generate_kb_activity_strength(self):
        """活动力度问答"""
        return f"我们这次是{self.brand}限时专享活动，给到您的也是很优惠的价格了，建议您可以到我们{self.store}对比看下哦"
    
    def _generate_kb_other_benefits(self):
        """其他福利问答"""
        benefits_text = ""
        if "加赠" in self.benefits:
            add_gift_text = self.benefits['加赠']
            if '宠物单笔实付满239元加赠' in add_gift_text:
                add_gift_text = add_gift_text.replace('宠物单笔实付满239元加赠', '满239元还')
            benefits_text = f"，{add_gift_text}"
        
        return f"除了优惠价格和礼赠{benefits_text}，那更多活动，您可以到{self.store}联系店铺客服详细了解一下呢~"
    
    def _clean_all_residuals(self, text):
        """清理所有原模板残留"""
        residuals = [
            "网易严选紫金烘焙猫粮",
            "紫金烘焙猫粮",
            "83元",
            "9件好礼",
            "猫条",
            "同款试吃",
            "正装宠物香氛",
            "50元专属优惠券",
            "一张50元",
            "猫条等9件好礼",
            "同款试吃装",
            "低至83元每包",
            "低至83元每袋",
        ]
        
        for residual in residuals:
            text = text.replace(residual, "")
        
        # 修复重复表达
        text = text.replace("的京东，", "的京东")
        text = text.replace("的京东。", "的京东")
        text = text.replace("一张，", "一张")
        
        return text
    
    def _extract_gift_count(self, gift_text):
        """提取赠品数量"""
        import re
        match = re.search(r'(\d+)件', gift_text)
        if match:
            return match.group(1)
        return gift_text

def create_final_script(scenario_name, scenario_info, template_wb):
    """创建最终版话术"""
    new_wb = openpyxl.load_workbook(TEMPLATE_FILE)
    
    ws_main = new_wb["主流程"]
    ws_kb = new_wb["知识库"]
    
    generator = FinalScriptGenerator(scenario_info)
    
    # 更新主流程
    for row_idx, row in enumerate(ws_main.iter_rows(min_row=2), start=2):
        node_name = str(row[1].value) if row[1].value else ""
        
        if node_name == "开场白":
            ws_main.cell(row=row_idx, column=3).value = generator.generate开场白()
        elif row_idx == 7:
            ws_main.cell(row=row_idx, column=3).value = generator.generate钩子()
        elif node_name == "用户忙":
            ws_main.cell(row=row_idx, column=3).value = generator.generate用户忙()
        elif "重述" in node_name:
            ws_main.cell(row=row_idx, column=3).value = generator.generate重述()
        elif node_name == "邀约到店-起始节点":
            ws_main.cell(row=row_idx, column=3).value = generator.generate邀约()
        elif node_name == "普通节点":
            ws_main.cell(row=row_idx, column=3).value = generator.generate挽回()
    
    # 更新知识库
    ws_kb = generator.update_knowledge_base(ws_kb)
    
    # 保存文件
    output_file = os.path.join(OUTPUT_DIR, f"{scenario_name}_话术.xlsx")
    new_wb.save(output_file)
    
    return output_file

def main():
    """主函数"""
    import re
    
    with open(SCENARIOS_FILE, 'r', encoding='utf-8') as f:
        scenarios = json.load(f)
    
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    print(f"开始生成最终版话术，修复所有质量问题...")
    print(f"场景数: {len(scenarios)}\n")
    
    output_files = []
    for scenario_name, scenario_info in scenarios.items():
        print(f"正在处理: {scenario_name}")
        template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
        output_file = create_final_script(scenario_name, scenario_info, template_wb)
        output_files.append(output_file)
        print(f"✓ 已生成: {output_file}\n")
    
    print(f"完成!共生成{len(output_files)}个最终版话术文件")
    print(f"输出目录: {OUTPUT_DIR}")
    print(f"\n下一步: 运行质量检查工具验证修复效果")

if __name__ == "__main__":
    main()